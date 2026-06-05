from django import forms
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.http import HttpResponse

# Imports para exportacion de horarios
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apphorarios.models import Asignatura, Aula, Clase, Estudiante, Grado, Hora, Profesor, UsuarioPersonalizado


ADMIN_DENIED_MESSAGE = 'No tienes permisos para acceder a esta pagina.'


def _guard_admin(request):
    if request.user.rol != 'admin':
        messages.error(request, ADMIN_DENIED_MESSAGE)
        return redirect('login')
    return None


class BootstrapFormMixin:
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for field in self.fields.values():
            widget = field.widget
            existing_class = widget.attrs.get('class', '')

            if isinstance(widget, (forms.Select, forms.SelectMultiple)):
                css_class = 'form-select'
            elif isinstance(widget, forms.CheckboxInput):
                css_class = 'form-check-input'
            else:
                css_class = 'form-control'

            widget.attrs['class'] = f'{existing_class} {css_class}'.strip()


class LoginForm(AuthenticationForm):
    TIPOS_USUARIO = [
        ('admin', 'Administrador'),
        ('estudiante', 'Estudiante'),
        ('profesor', 'Profesor'),
    ]

    tipo_usuario = forms.ChoiceField(
        choices=TIPOS_USUARIO,
        label='Tipo de usuario',
        widget=forms.Select(),
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['username'].widget.attrs.update(
            {'class': 'form-control', 'placeholder': 'Usuario o codigo'}
        )
        self.fields['password'].widget.attrs.update(
            {'class': 'form-control', 'placeholder': 'Contrasena'}
        )
        self.fields['tipo_usuario'].widget.attrs.update({'class': 'form-select'})


class EstudianteForm(BootstrapFormMixin, forms.ModelForm):
    class Meta:
        model = Estudiante
        fields = [
            'codigo_estudiante',
            'identificacion',
            'primer_nombre',
            'segundo_nombre',
            'primer_apellido',
            'segundo_apellido',
            'grado',
        ]
        labels = {
            'codigo_estudiante': 'Codigo del estudiante',
            'identificacion': 'Identificacion',
            'primer_nombre': 'Primer nombre',
            'segundo_nombre': 'Segundo nombre',
            'primer_apellido': 'Primer apellido',
            'segundo_apellido': 'Segundo apellido',
            'grado': 'Grado',
        }


class ProfesorForm(BootstrapFormMixin, forms.ModelForm):
    class Meta:
        model = Profesor
        fields = [
            'codigo_profesor',
            'identificacion',
            'primer_nombre',
            'segundo_nombre',
            'primer_apellido',
            'segundo_apellido',
            'asignatura',
            'grados',
        ]
        labels = {
            'codigo_profesor': 'Codigo del profesor',
            'identificacion': 'Identificacion',
            'primer_nombre': 'Primer nombre',
            'segundo_nombre': 'Segundo nombre',
            'primer_apellido': 'Primer apellido',
            'segundo_apellido': 'Segundo apellido',
            'asignatura': 'Asignatura principal',
            'grados': 'Grados asignados',
        }
        widgets = {
            'grados': forms.SelectMultiple(attrs={'size': 6}),
        }


class AulaForm(BootstrapFormMixin, forms.ModelForm):
    class Meta:
        model = Aula
        fields = ['nombre', 'capacidad']
        labels = {'nombre': 'Nombre del aula', 'capacidad': 'Capacidad'}


class AsignaturaForm(BootstrapFormMixin, forms.ModelForm):
    class Meta:
        model = Asignatura
        fields = ['codigo_asignatura', 'nombre']
        labels = {'codigo_asignatura': 'Codigo', 'nombre': 'Nombre'}


class GradoForm(BootstrapFormMixin, forms.ModelForm):
    class Meta:
        model = Grado
        fields = ['codigo_grado', 'nombre']
        labels = {'codigo_grado': 'Codigo', 'nombre': 'Nombre del grado'}


class HoraForm(BootstrapFormMixin, forms.ModelForm):
    class Meta:
        model = Hora
        fields = ['hora_inicio', 'hora_fin']
        labels = {'hora_inicio': 'Hora de inicio', 'hora_fin': 'Hora de fin'}
        widgets = {
            'hora_inicio': forms.TimeInput(attrs={'type': 'time'}),
            'hora_fin': forms.TimeInput(attrs={'type': 'time'}),
        }


class ClaseForm(BootstrapFormMixin, forms.ModelForm):
    class Meta:
        model = Clase
        fields = ['descripcion_clase', 'profesor', 'grado', 'aula', 'hora', 'dia']
        labels = {
            'descripcion_clase': 'Nombre o descripcion',
            'profesor': 'Profesor',
            'grado': 'Grado',
            'aula': 'Aula',
            'hora': 'Bloque horario',
            'dia': 'Dia',
        }
        help_texts = {
            'grado': 'Solo debe existir una clase por grado en el mismo bloque horario.',
        }


def _sync_student_user(estudiante):
    user = estudiante.usuario

    if user is None:
        existing_user = UsuarioPersonalizado.objects.filter(
            username=estudiante.codigo_estudiante
        ).first()
        if existing_user:
            has_other_student = hasattr(existing_user, 'estudiante_perfil') and existing_user.estudiante_perfil != estudiante
            has_professor_profile = hasattr(existing_user, 'profesor_perfil')
            if has_other_student or has_professor_profile:
                raise ValueError('Ya existe otro usuario vinculado con este codigo de estudiante.')
        user = existing_user

    if user is None:
        user = UsuarioPersonalizado.objects.create_user(
            username=estudiante.codigo_estudiante,
            password=estudiante.identificacion,
            rol='estudiante',
        )
    else:
        user.username = estudiante.codigo_estudiante
        user.rol = 'estudiante'
        user.set_password(estudiante.identificacion)
        user.save()

    estudiante.usuario = user


def _sync_professor_user(profesor):
    user = profesor.usuario

    if user is None:
        existing_user = UsuarioPersonalizado.objects.filter(
            username=profesor.codigo_profesor
        ).first()
        if existing_user:
            has_other_professor = hasattr(existing_user, 'profesor_perfil') and existing_user.profesor_perfil != profesor
            has_student_profile = hasattr(existing_user, 'estudiante_perfil')
            if has_other_professor or has_student_profile:
                raise ValueError('Ya existe otro usuario vinculado con este codigo de profesor.')
        user = existing_user

    if user is None:
        user = UsuarioPersonalizado.objects.create_user(
            username=profesor.codigo_profesor,
            password=profesor.identificacion,
            rol='profesor',
        )
    else:
        user.username = profesor.codigo_profesor
        user.rol = 'profesor'
        user.set_password(profesor.identificacion)
        user.save()

    profesor.usuario = user


def _delete_linked_user(instance):
    user = getattr(instance, 'usuario', None)
    instance.delete()
    if user:
        user.delete()


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    form = LoginForm(request, data=request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            tipo_usuario = form.cleaned_data.get('tipo_usuario')
            user = authenticate(username=username, password=password)

            if user is not None and user.rol == tipo_usuario:
                login(request, user)
                messages.success(request, f'Bienvenido, {user.username}.')
                return redirect('admin_dashboard' if user.rol == 'admin' else 'horario_view')

        messages.error(request, 'Las credenciales o el tipo de usuario no son correctos.')

    return render(request, 'login.html', {'form': form})


def dashboard_view(request):
    if request.user.rol == 'admin':
        return redirect('admin_dashboard')
    return redirect('horario_view')


@login_required
def _obtener_datos_horario(request, requested_user_code, requested_user_type):
    viewed_user = request.user
    viewed_role = request.user.rol
    viewed_label = request.user.username
    viewed_code = request.user.username

    if request.user.rol == 'admin' and requested_user_code and requested_user_type:
        if requested_user_type == 'profesor':
            perfil = get_object_or_404(
                Profesor.objects.select_related('usuario', 'asignatura'),
                codigo_profesor=requested_user_code,
            )
            viewed_user = perfil.usuario
            viewed_role = 'profesor'
            viewed_label = f'{perfil.primer_nombre} {perfil.primer_apellido}'
            viewed_code = perfil.codigo_profesor
        elif requested_user_type == 'estudiante':
            perfil = get_object_or_404(
                Estudiante.objects.select_related('usuario', 'grado'),
                codigo_estudiante=requested_user_code,
            )
            viewed_user = perfil.usuario
            viewed_role = 'estudiante'
            viewed_label = f'{perfil.primer_nombre} {perfil.primer_apellido}'
            viewed_code = perfil.codigo_estudiante
        elif requested_user_type == 'grado':
            perfil = get_object_or_404(
                Grado,
                codigo_grado=requested_user_code,
            )
            viewed_user = None
            viewed_role = 'grado'
            viewed_label = f'{perfil.codigo_grado} - {perfil.nombre}'
            viewed_code = perfil.codigo_grado

    day_labels = {
        'Lunes': [],
        'Martes': [],
        'Miercoles': [],
        'Jueves': [],
        'Viernes': [],
    }
    day_by_code = {
        'LU': 'Lunes',
        'MA': 'Martes',
        'MI': 'Miercoles',
        'JU': 'Jueves',
        'VI': 'Viernes',
    }
    horario_por_dia = day_labels
    total_clases = 0

    if viewed_role == 'profesor':
        profesor = getattr(viewed_user, 'profesor_perfil', None)
        if profesor is None:
            raise ValueError('No fue posible encontrar el perfil del profesor.')
        viewed_label = f'{profesor.primer_nombre} {profesor.primer_apellido}'
        viewed_code = profesor.codigo_profesor

        clases = (
            Clase.objects.filter(profesor=profesor)
            .select_related('hora', 'aula', 'grado', 'profesor__asignatura')
            .order_by('dia', 'hora__hora_inicio')
        )

        for clase in clases:
            dia_str = day_by_code.get(clase.dia)
            horario_por_dia[dia_str].append(
                {
                    'hora': str(clase.hora),
                    'titulo': clase.descripcion_clase,
                    'detalle_principal': clase.grado.nombre,
                    'detalle_secundario': clase.profesor.asignatura.nombre,
                    'aula': clase.aula.nombre,
                }
            )
        total_clases = clases.count()
    elif viewed_role == 'estudiante':
        estudiante = getattr(viewed_user, 'estudiante_perfil', None)
        if estudiante is None:
            raise ValueError('No fue posible encontrar el perfil del estudiante.')
        viewed_label = f'{estudiante.primer_nombre} {estudiante.primer_apellido}'
        viewed_code = estudiante.codigo_estudiante

        clases = (
            Clase.objects.filter(grado=estudiante.grado)
            .select_related('hora', 'aula', 'profesor__asignatura')
            .order_by('dia', 'hora__hora_inicio')
        )

        for clase in clases:
            dia_str = day_by_code.get(clase.dia)
            horario_por_dia[dia_str].append(
                {
                    'hora': str(clase.hora),
                    'titulo': clase.profesor.asignatura.nombre,
                    'detalle_principal': f'{clase.profesor.primer_nombre} {clase.profesor.primer_apellido}',
                    'detalle_secundario': clase.descripcion_clase,
                    'aula': clase.aula.nombre,
                }
            )
        total_clases = clases.count()
    elif viewed_role == 'grado':
        grado = get_object_or_404(Grado, codigo_grado=viewed_code)
        clases = (
            Clase.objects.filter(grado=grado)
            .select_related('hora', 'aula', 'profesor__asignatura')
            .order_by('dia', 'hora__hora_inicio')
        )

        for clase in clases:
            dia_str = day_by_code.get(clase.dia)
            horario_por_dia[dia_str].append(
                {
                    'hora': str(clase.hora),
                    'titulo': clase.profesor.asignatura.nombre,
                    'detalle_principal': f'{clase.profesor.primer_nombre} {clase.profesor.primer_apellido}',
                    'detalle_secundario': clase.descripcion_clase,
                    'aula': clase.aula.nombre,
                }
            )
        total_clases = clases.count()
    else:
        raise ValueError('Solo estudiantes, profesores o administradores pueden visualizar horarios.')

    return {
        'horario_por_dia': horario_por_dia,
        'user_rol': viewed_role,
        'viewed_label': viewed_label,
        'viewed_code': viewed_code,
        'total_clases': total_clases,
        'dias_activos': sum(1 for clases_dia in horario_por_dia.values() if clases_dia),
    }


@login_required
def horario_view(request):
    requested_user_code = request.GET.get('user_code')
    requested_user_type = request.GET.get('user_type')

    if request.user.rol == 'admin' and not (requested_user_code and requested_user_type):
        messages.info(request, 'Selecciona primero a quien deseas consultar.')
        return redirect('ver_horarios')

    try:
        data = _obtener_datos_horario(request, requested_user_code, requested_user_type)
    except Exception as e:
        messages.error(request, str(e))
        return redirect('dashboard')

    context = {
        'horario_por_dia': data['horario_por_dia'],
        'user_rol': data['user_rol'],
        'viewed_label': data['viewed_label'],
        'viewed_code': data['viewed_code'],
        'total_clases': data['total_clases'],
        'dias_activos': data['dias_activos'],
        'admin_view': request.user.rol == 'admin',
        'requested_user_code': requested_user_code or '',
        'requested_user_type': requested_user_type or '',
    }
    return render(request, 'horarios/horario_view.html', context)


@login_required
def exportar_horario_pdf(request):
    requested_user_code = request.GET.get('user_code')
    requested_user_type = request.GET.get('user_type')

    if request.user.rol != 'admin':
        requested_user_code = request.user.username
        requested_user_type = request.user.rol

    try:
        data = _obtener_datos_horario(request, requested_user_code, requested_user_type)
    except Exception as e:
        messages.error(request, str(e))
        return redirect('dashboard')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="horario_{data["viewed_code"]}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.HexColor('#0f5c5c'),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        textColor=colors.HexColor('#5b6470'),
        spaceAfter=15
    )
    day_title_style = ParagraphStyle(
        'DayTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#0f5c5c'),
        alignment=1
    )
    class_title_style = ParagraphStyle(
        'ClassTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.HexColor('#1f2a37')
    )
    class_detail_style = ParagraphStyle(
        'ClassDetailStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#5b6470')
    )

    story.append(Paragraph(f"EduSchedule - Horario Semanal", title_style))
    story.append(Paragraph(f"{data['viewed_label']} ({data['user_rol'].upper()}: {data['viewed_code']})", subtitle_style))

    headers = [Paragraph(f"<b>Lunes</b>", day_title_style),
               Paragraph(f"<b>Martes</b>", day_title_style),
               Paragraph(f"<b>Miércoles</b>", day_title_style),
               Paragraph(f"<b>Jueves</b>", day_title_style),
               Paragraph(f"<b>Viernes</b>", day_title_style)]

    horario_por_dia = data['horario_por_dia']
    max_rows = max(len(horario_por_dia[dia]) for dia in ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes'])

    table_data = [headers]

    for row_idx in range(max_rows):
        row = []
        for dia in ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']:
            clases_dia = horario_por_dia[dia]
            if row_idx < len(clases_dia):
                clase = clases_dia[row_idx]
                cell_elements = [
                    Paragraph(f"<b>{clase['hora']}</b>", class_detail_style),
                    Paragraph(clase['titulo'], class_title_style),
                    Paragraph(f"{clase['detalle_principal']}", class_detail_style),
                    Paragraph(f"Aula: {clase['aula']}", class_detail_style)
                ]
                row.append(cell_elements)
            else:
                row.append(Paragraph("", class_detail_style))
        table_data.append(row)

    t = Table(table_data, colWidths=[144, 144, 144, 144, 144])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e6f0f0')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#0f5c5c')),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#fafafa')),
    ]))

    story.append(t)
    doc.build(story)
    return response


@login_required
def exportar_horario_excel(request):
    requested_user_code = request.GET.get('user_code')
    requested_user_type = request.GET.get('user_type')

    if request.user.rol != 'admin':
        requested_user_code = request.user.username
        requested_user_type = request.user.rol

    try:
        data = _obtener_datos_horario(request, requested_user_code, requested_user_type)
    except Exception as e:
        messages.error(request, str(e))
        return redirect('dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horario Semanal"

    ws.views.sheetView[0].showGridLines = True

    primary_color = "0F5C5C"
    light_bg = "FAF7EF"

    ws['A1'] = "EduSchedule - Horario Semanal"
    ws['A1'].font = Font(name="Calibri", size=16, bold=True, color="0F5C5C")
    ws.merge_cells('A1:E1')

    ws['A2'] = f"{data['viewed_label']} ({data['user_rol'].upper()}: {data['viewed_code']})"
    ws['A2'].font = Font(name="Calibri", size=11, italic=True)
    ws.merge_cells('A2:E2')

    ws.append([])

    headers = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    ws.append(headers)

    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, 6):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    horario_por_dia = data['horario_por_dia']
    max_rows = max(len(horario_por_dia[dia]) for dia in ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes'])

    thin_border = Border(
        left=Side(style='thin', color="DDDDDD"),
        right=Side(style='thin', color="DDDDDD"),
        top=Side(style='thin', color="DDDDDD"),
        bottom=Side(style='thin', color="DDDDDD")
    )

    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row_idx in range(max_rows):
        excel_row_num = 5 + row_idx
        for col_idx, dia in enumerate(['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes'], start=1):
            clases_dia = horario_por_dia[dia]
            cell = ws.cell(row=excel_row_num, column=col_idx)
            cell.alignment = cell_align
            cell.border = thin_border
            
            if row_idx < len(clases_dia):
                clase = clases_dia[row_idx]
                text = f"[{clase['hora']}]\n{clase['titulo']}\n{clase['detalle_principal']}\nAula: {clase['aula']}"
                cell.value = text
                cell.fill = PatternFill(start_color=light_bg, end_color=light_bg, fill_type="solid")
            else:
                cell.value = ""

    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 25

    ws.row_dimensions[4].height = 25
    for row_idx in range(max_rows):
        ws.row_dimensions[5 + row_idx].height = 75

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="horario_{data["viewed_code"]}.xlsx"'
    wb.save(response)
    return response


@login_required
def admin_dashboard(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    context = {
        'year': '2026',
        'period': '1',
        'institution_name': 'Institucion Educativa Demo',
        'metrics': [
            {'label': 'Estudiantes', 'value': Estudiante.objects.count(), 'icon': 'fa-user-graduate'},
            {'label': 'Profesores', 'value': Profesor.objects.count(), 'icon': 'fa-chalkboard-user'},
            {'label': 'Aulas', 'value': Aula.objects.count(), 'icon': 'fa-door-open'},
            {'label': 'Clases', 'value': Clase.objects.count(), 'icon': 'fa-calendar-days'},
        ],
        'pending_items': [
            {
                'label': 'Profesores sin grados asignados',
                'value': Profesor.objects.filter(grados__isnull=True).distinct().count(),
            },
            {
                'label': 'Estudiantes sin usuario vinculado',
                'value': Estudiante.objects.filter(usuario__isnull=True).count(),
            },
            {
                'label': 'Profesores sin usuario vinculado',
                'value': Profesor.objects.filter(usuario__isnull=True).count(),
            },
        ],
    }
    return render(request, 'admin/dashboard.html', context)


@login_required
def lista_estudiantes(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect('gestionar_estudiantes')


@login_required
def lista_profesores(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect('gestionar_profesores')


@login_required
def lista_aulas(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect('gestionar_aulas')


@login_required
def lista_clases(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect('gestionar_clases')


@login_required
def lista_asignaturas(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect('gestionar_asignaturas')


@login_required
def lista_horas(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect('gestionar_horas')


@login_required
def lista_grados(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect
    return redirect('gestionar_grados')


@login_required
def editar_estudiante(request, estudiante_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    estudiante = get_object_or_404(Estudiante, id=estudiante_id)
    form = EstudianteForm(request.POST or None, instance=estudiante)

    if request.method == 'POST' and form.is_valid():
        estudiante = form.save(commit=False)
        try:
            _sync_student_user(estudiante)
        except ValueError as error:
            form.add_error('codigo_estudiante', str(error))
        else:
            estudiante.save()
            messages.success(request, 'Estudiante actualizado correctamente.')
            return redirect('gestionar_estudiantes')

    return render(request, 'admin/editar_estudiante.html', {'form': form, 'estudiante': estudiante})


@login_required
def gestionar_estudiantes(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    query = request.GET.get('q', '').strip()
    estudiantes = Estudiante.objects.select_related('grado', 'usuario').order_by('codigo_estudiante')
    if query:
        estudiantes = estudiantes.filter(
            Q(primer_nombre__icontains=query)
            | Q(primer_apellido__icontains=query)
            | Q(codigo_estudiante__icontains=query)
        )

    form = EstudianteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        estudiante = form.save(commit=False)
        try:
            _sync_student_user(estudiante)
        except ValueError as error:
            form.add_error('codigo_estudiante', str(error))
        else:
            estudiante.save()
            messages.success(request, 'Estudiante agregado correctamente.')
            return redirect('gestionar_estudiantes')

    context = {'estudiantes': estudiantes, 'form': form, 'query': query}
    return render(request, 'admin/gestionar_estudiantes.html', context)


@login_required
def eliminar_estudiante(request, estudiante_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    estudiante = get_object_or_404(Estudiante, id=estudiante_id)
    _delete_linked_user(estudiante)
    messages.success(request, 'Estudiante eliminado correctamente.')
    return redirect('gestionar_estudiantes')


@login_required
def editar_profesor(request, profesor_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    profesor = get_object_or_404(Profesor, id=profesor_id)
    form = ProfesorForm(request.POST or None, instance=profesor)

    if request.method == 'POST' and form.is_valid():
        profesor = form.save(commit=False)
        try:
            _sync_professor_user(profesor)
        except ValueError as error:
            form.add_error('codigo_profesor', str(error))
        else:
            profesor.save()
            form.save_m2m()
            messages.success(request, 'Profesor actualizado correctamente.')
            return redirect('gestionar_profesores')

    return render(request, 'admin/editar_profesor.html', {'form': form, 'profesor': profesor})


@login_required
def gestionar_profesores(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    query = request.GET.get('q', '').strip()
    profesores = (
        Profesor.objects.select_related('asignatura', 'usuario')
        .prefetch_related('grados')
        .order_by('codigo_profesor')
    )
    if query:
        profesores = profesores.filter(
            Q(primer_nombre__icontains=query)
            | Q(primer_apellido__icontains=query)
            | Q(codigo_profesor__icontains=query)
        )

    form = ProfesorForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        profesor = form.save(commit=False)
        try:
            _sync_professor_user(profesor)
        except ValueError as error:
            form.add_error('codigo_profesor', str(error))
        else:
            profesor.save()
            form.save_m2m()
            messages.success(request, 'Profesor agregado correctamente.')
            return redirect('gestionar_profesores')

    context = {'profesores': profesores, 'form': form, 'query': query}
    return render(request, 'admin/gestionar_profesores.html', context)


@login_required
def eliminar_profesor(request, profesor_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    profesor = get_object_or_404(Profesor, id=profesor_id)
    _delete_linked_user(profesor)
    messages.success(request, 'Profesor eliminado correctamente.')
    return redirect('gestionar_profesores')


@login_required
def editar_aula(request, aula_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    aula = get_object_or_404(Aula, id=aula_id)
    form = AulaForm(request.POST or None, instance=aula)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Aula actualizada correctamente.')
        return redirect('gestionar_aulas')
    return render(request, 'admin/editar_aula.html', {'form': form, 'aula': aula})


@login_required
def gestionar_aulas(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    query = request.GET.get('q', '').strip()
    aulas = Aula.objects.order_by('nombre')
    if query:
        aulas = aulas.filter(nombre__icontains=query)

    form = AulaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Aula agregada correctamente.')
        return redirect('gestionar_aulas')

    context = {'aulas': aulas, 'form': form, 'query': query}
    return render(request, 'admin/gestionar_aulas.html', context)


@login_required
def eliminar_aula(request, aula_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    aula = get_object_or_404(Aula, id=aula_id)
    aula.delete()
    messages.success(request, 'Aula eliminada correctamente.')
    return redirect('gestionar_aulas')


@login_required
def editar_asignatura(request, asignatura_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    asignatura = get_object_or_404(Asignatura, id=asignatura_id)
    form = AsignaturaForm(request.POST or None, instance=asignatura)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Asignatura actualizada correctamente.')
        return redirect('gestionar_asignaturas')
    return render(request, 'admin/editar_asignatura.html', {'form': form, 'asignatura': asignatura})


@login_required
def gestionar_asignaturas(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    query = request.GET.get('q', '').strip()
    asignaturas = Asignatura.objects.order_by('codigo_asignatura')
    if query:
        asignaturas = asignaturas.filter(
            Q(nombre__icontains=query) | Q(codigo_asignatura__icontains=query)
        )

    form = AsignaturaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Asignatura agregada correctamente.')
        return redirect('gestionar_asignaturas')

    context = {'asignaturas': asignaturas, 'form': form, 'query': query}
    return render(request, 'admin/gestionar_asignaturas.html', context)


@login_required
def eliminar_asignatura(request, asignatura_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    asignatura = get_object_or_404(Asignatura, id=asignatura_id)
    asignatura.delete()
    messages.success(request, 'Asignatura eliminada correctamente.')
    return redirect('gestionar_asignaturas')


@login_required
def editar_grado(request, grado_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    grado = get_object_or_404(Grado, id=grado_id)
    form = GradoForm(request.POST or None, instance=grado)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Grado actualizado correctamente.')
        return redirect('gestionar_grados')
    return render(request, 'admin/editar_grado.html', {'form': form, 'grado': grado})


@login_required
def gestionar_grados(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    query = request.GET.get('q', '').strip()
    grados = Grado.objects.order_by('codigo_grado')
    if query:
        grados = grados.filter(
            Q(nombre__icontains=query) | Q(codigo_grado__icontains=query)
        )

    form = GradoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Grado agregado correctamente.')
        return redirect('gestionar_grados')

    context = {'grados': grados, 'form': form, 'query': query}
    return render(request, 'admin/gestionar_grados.html', context)


@login_required
def eliminar_grado(request, grado_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    grado = get_object_or_404(Grado, id=grado_id)
    grado.delete()
    messages.success(request, 'Grado eliminado correctamente.')
    return redirect('gestionar_grados')


@login_required
def editar_hora(request, hora_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    hora = get_object_or_404(Hora, id=hora_id)
    form = HoraForm(request.POST or None, instance=hora)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Bloque horario actualizado correctamente.')
        return redirect('gestionar_horas')
    return render(request, 'admin/editar_hora.html', {'form': form, 'hora': hora})


@login_required
def gestionar_horas(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    query = request.GET.get('q', '').strip()
    horas = Hora.objects.order_by('hora_inicio')
    if query:
        horas = horas.filter(
            Q(hora_inicio__icontains=query) | Q(hora_fin__icontains=query)
        )

    form = HoraForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Bloque horario agregado correctamente.')
        return redirect('gestionar_horas')

    context = {'horas': horas, 'form': form, 'query': query}
    return render(request, 'admin/gestionar_horas.html', context)


@login_required
def eliminar_hora(request, hora_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    hora = get_object_or_404(Hora, id=hora_id)
    hora.delete()
    messages.success(request, 'Bloque horario eliminado correctamente.')
    return redirect('gestionar_horas')


@login_required
def gestionar_clases(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    query = request.GET.get('q', '').strip()
    clases = (
        Clase.objects.select_related('profesor__asignatura', 'aula', 'hora', 'grado')
        .prefetch_related('grado__estudiantes')
        .order_by('dia', 'hora__hora_inicio', 'grado__codigo_grado')
    )
    if query:
        clases = clases.filter(
            Q(descripcion_clase__icontains=query)
            | Q(grado__nombre__icontains=query)
            | Q(profesor__primer_nombre__icontains=query)
            | Q(profesor__primer_apellido__icontains=query)
        )

    form = ClaseForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Clase programada correctamente.')
        return redirect('gestionar_clases')

    context = {'clases': clases, 'form': form, 'query': query}
    return render(request, 'admin/gestionar_clases.html', context)


@login_required
def editar_clase(request, clase_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    clase = get_object_or_404(Clase, id=clase_id)
    form = ClaseForm(request.POST or None, instance=clase)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Clase actualizada correctamente.')
        return redirect('gestionar_clases')

    return render(request, 'admin/editar_clase.html', {'form': form, 'clase': clase})


@login_required
def eliminar_clase(request, clase_id):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    clase = get_object_or_404(Clase, id=clase_id)
    clase.delete()
    messages.success(request, 'Clase eliminada correctamente.')
    return redirect('gestionar_clases')


@login_required
def ver_horarios(request):
    admin_redirect = _guard_admin(request)
    if admin_redirect:
        return admin_redirect

    selected_user_code = request.GET.get('user_code', '')
    selected_user_type = request.GET.get('user_type', '')

    if selected_user_code and selected_user_type:
        return redirect(
            reverse('horario_view')
            + f'?user_code={selected_user_code}&user_type={selected_user_type}'
        )

    context = {
        'message': 'Selecciona un profesor, estudiante o grado para consultar su horario.',
        'selected_user_type': selected_user_type,
        'profesores': Profesor.objects.select_related('asignatura').order_by('primer_nombre', 'primer_apellido'),
        'estudiantes': Estudiante.objects.select_related('grado').order_by('primer_nombre', 'primer_apellido'),
        'grados': Grado.objects.all().order_by('codigo_grado'),
    }
    return render(request, 'admin/ver_horarios.html', context)
